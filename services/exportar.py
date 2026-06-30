"""
exportar.py - Generación de reporte PDF profesional, exportación a Excel
y guardado de la gráfica como imagen.
"""
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Image,
    Table,
    TableStyle,
    HRFlowable,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

from tkinter import filedialog, messagebox
from datetime import datetime
import os
import sys
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from database.database import obtener_datos, obtener_ultimo_registro


def _resource_path(ruta_relativa):
    """
    Devuelve la ruta absoluta a un archivo incluido con la app (como el logo),
    funcionando tanto si se ejecuta como script normal como si está
    empaquetado con PyInstaller.

    PyInstaller, cuando usa --add-data, extrae esos archivos a una carpeta
    temporal accesible en sys._MEIPASS (solo existe en modo empaquetado).
    """
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        base_path = sys._MEIPASS
    else:
        # Carpeta raíz del proyecto (un nivel arriba de services/)
        base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, ruta_relativa)

# Colores institucionales (ajusta a tu paleta real si tienes manual de marca)
COLOR_PRIMARIO = colors.HexColor("#1B3A2A")   # verde oscuro institucional
COLOR_SECUNDARIO = colors.HexColor("#4E8A1E")
COLOR_TEXTO_SUAVE = colors.HexColor("#555555")
COLOR_FILA_PAR = colors.HexColor("#F2F4F1")


def _generar_imagen_grafica(red, fig_existente=None):
    """
    Devuelve un BytesIO con la gráfica en PNG, lista para insertar en el PDF.
    Si se pasa una figura de matplotlib ya dibujada (fig_existente), la reutiliza.
    Si no, no se incluye gráfica (se omite esa sección).
    """
    if fig_existente is None:
        return None
    buf = io.BytesIO()
    fig_existente.savefig(buf, format="png", dpi=150, bbox_inches="tight",
                           facecolor=fig_existente.get_facecolor())
    buf.seek(0)
    return buf


def _calcular_crecimiento(datos):
    """Calcula crecimiento absoluto y porcentual entre el primer y último registro."""
    if len(datos) < 2:
        return None

    primero = datos[0]
    ultimo = datos[-1]
    resultado = {}
    for carrera in ("Agropecuaria", "Agronegocios", "Agroindustrial"):
        inicial = primero[carrera] or 0
        final = ultimo[carrera] or 0
        diferencia = final - inicial
        porcentaje = (diferencia / inicial * 100) if inicial > 0 else 0
        resultado[carrera] = {
            "inicial": inicial,
            "final": final,
            "diferencia": diferencia,
            "porcentaje": porcentaje,
        }
    return resultado


def generar_pdf(red, fig_grafica=None):
    """
    Genera un reporte PDF profesional para la red social indicada.

    Parámetros:
        red: nombre de la red social ("Facebook", "Instagram", "TikTok")
        fig_grafica: (opcional) objeto Figure de matplotlib ya dibujado,
                     para incrustar la gráfica de evolución en el PDF.
    """
    archivo = filedialog.asksaveasfilename(
        defaultextension=".pdf",
        filetypes=[("PDF", "*.pdf")],
        initialfile=f"Reporte_{red}_{datetime.now().strftime('%Y%m%d')}.pdf",
    )
    if not archivo:
        return

    doc = SimpleDocTemplate(
        archivo,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=50,
        bottomMargin=60,
    )

    estilos = getSampleStyleSheet()
    estilos.add(ParagraphStyle(
        name="Subtitulo", parent=estilos["Normal"],
        fontSize=11, textColor=COLOR_TEXTO_SUAVE, alignment=TA_CENTER,
        spaceAfter=4,
    ))
    estilos.add(ParagraphStyle(
        name="FechaGeneracion", parent=estilos["Normal"],
        fontSize=9, textColor=COLOR_TEXTO_SUAVE, alignment=TA_RIGHT,
    ))
    estilos.add(ParagraphStyle(
        name="SeccionTitulo", parent=estilos["Heading2"],
        textColor=COLOR_PRIMARIO, spaceBefore=14, spaceAfter=8,
    ))

    elementos = []

    # ───────────────── ENCABEZADO ─────────────────
    logo_path = _resource_path(os.path.join("imagenes", "logo_uleam.png"))
    if os.path.exists(logo_path):
        elementos.append(Image(logo_path, width=140, height=70))
        elementos.append(Spacer(1, 10))

    elementos.append(Paragraph(
        "REPORTE DE MONITOREO DE REDES SOCIALES", estilos["Title"]
    ))
    elementos.append(Paragraph(
        "Facultad de Ciencias Veterinarias y Agropecuarias — ULEAM",
        estilos["Subtitulo"]
    ))
    elementos.append(Paragraph(
        f"Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}",
        estilos["FechaGeneracion"]
    ))
    elementos.append(Spacer(1, 6))
    elementos.append(HRFlowable(width="100%", thickness=1.2, color=COLOR_PRIMARIO))
    elementos.append(Spacer(1, 16))

    # ───────────────── RED SELECCIONADA ─────────────────
    elementos.append(Paragraph(f"Red social: <b>{red}</b>", estilos["SeccionTitulo"]))

    datos_historicos = obtener_datos(red)
    ultimo = obtener_ultimo_registro(red)

    if not datos_historicos or not ultimo:
        elementos.append(Paragraph(
            "No hay registros disponibles para esta red social.", estilos["Normal"]
        ))
        doc.build(elementos)
        messagebox.showinfo("PDF", "Reporte generado (sin datos disponibles).")
        return

    # ───────────────── RESUMEN ACTUAL ─────────────────
    elementos.append(Paragraph("Estado actual", estilos["SeccionTitulo"]))
    tabla_actual = [
        ["Carrera", "Seguidores actuales"],
        ["Agropecuaria", f"{ultimo['Agropecuaria']:,}"],
        ["Agronegocios", f"{ultimo['Agronegocios']:,}"],
        ["Agroindustrial", f"{ultimo['Agroindustrial']:,}"],
    ]
    t1 = Table(tabla_actual, colWidths=[8 * cm, 6 * cm])
    t1.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), COLOR_PRIMARIO),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, COLOR_FILA_PAR]),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elementos.append(t1)
    elementos.append(Paragraph(
        f"<i>Último registro: {ultimo['fecha']}</i>", estilos["Subtitulo"]
    ))
    elementos.append(Spacer(1, 14))

    # ───────────────── CRECIMIENTO ─────────────────
    crecimiento = _calcular_crecimiento(datos_historicos)
    if crecimiento:
        elementos.append(Paragraph(
            f"Crecimiento desde el {datos_historicos[0]['fecha']}",
            estilos["SeccionTitulo"]
        ))
        tabla_crec = [["Carrera", "Inicial", "Actual", "Diferencia", "% Crecimiento"]]
        for carrera, vals in crecimiento.items():
            signo = "+" if vals["diferencia"] >= 0 else ""
            tabla_crec.append([
                carrera,
                f"{vals['inicial']:,}",
                f"{vals['final']:,}",
                f"{signo}{vals['diferencia']:,}",
                f"{signo}{vals['porcentaje']:.1f}%",
            ])
        t2 = Table(tabla_crec, colWidths=[4.2 * cm, 2.7 * cm, 2.7 * cm, 2.8 * cm, 2.8 * cm])
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), COLOR_SECUNDARIO),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, COLOR_FILA_PAR]),
            ("FONTSIZE", (0, 0), (-1, -1), 9.5),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elementos.append(t2)
        elementos.append(Spacer(1, 16))

    # ───────────────── GRÁFICA DE EVOLUCIÓN ─────────────────
    imagen_buf = _generar_imagen_grafica(red, fig_grafica)
    if imagen_buf:
        elementos.append(Paragraph("Evolución histórica", estilos["SeccionTitulo"]))
        elementos.append(Image(imagen_buf, width=17 * cm, height=9 * cm))
        elementos.append(Spacer(1, 14))

    # ───────────────── HISTÓRICO COMPLETO ─────────────────
    elementos.append(Paragraph("Histórico de registros", estilos["SeccionTitulo"]))
    tabla_hist = [["Fecha", "Agropecuaria", "Agronegocios", "Agroindustrial"]]
    for d in datos_historicos:
        tabla_hist.append([
            d["fecha"], f"{d['Agropecuaria']:,}",
            f"{d['Agronegocios']:,}", f"{d['Agroindustrial']:,}",
        ])
    t3 = Table(tabla_hist, colWidths=[5 * cm, 4 * cm, 4 * cm, 4 * cm], repeatRows=1)
    t3.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), COLOR_PRIMARIO),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, COLOR_FILA_PAR]),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elementos.append(t3)

    # ───────────────── PIE DE PÁGINA ─────────────────
    def pie_de_pagina(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.grey)
        canvas.line(40, 45, 570, 45)
        canvas.setFont("Helvetica", 8.5)
        canvas.setFillColor(COLOR_TEXTO_SUAVE)
        canvas.drawString(40, 30, "ULEAM — Sistema de Monitoreo de Redes Sociales")
        canvas.drawRightString(570, 30, f"Página {doc.page}")
        canvas.restoreState()

    doc.build(elementos, onFirstPage=pie_de_pagina, onLaterPages=pie_de_pagina)
    messagebox.showinfo("PDF", "Reporte profesional generado correctamente.")


# ──────────────────────────────────────────────────────────────────────────
# EXPORTAR A EXCEL
# ──────────────────────────────────────────────────────────────────────────
def exportar_excel(red=None):
    """
    Exporta el histórico de registros a un archivo .xlsx.
    Si se indica una red social, exporta solo esa; si no, exporta las tres
    en hojas separadas.
    """
    archivo = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")],
        initialfile=f"Seguidores_{red or 'Todas'}_{datetime.now().strftime('%Y%m%d')}.xlsx",
    )
    if not archivo:
        return

    redes = [red] if red else ["Facebook", "Instagram", "TikTok"]
    wb = Workbook()
    wb.remove(wb.active)  # quitamos la hoja vacía por defecto

    encabezado_fill = PatternFill(start_color="1B3A2A", end_color="1B3A2A", fill_type="solid")
    encabezado_font = Font(color="FFFFFF", bold=True)
    centrado = Alignment(horizontal="center")

    hojas_creadas = 0
    for r in redes:
        datos = obtener_datos(r)
        if not datos:
            continue

        ws = wb.create_sheet(title=r[:31])  # Excel limita el nombre de hoja a 31 caracteres
        encabezados = ["Fecha", "Agropecuaria", "Agronegocios", "Agroindustrial"]
        ws.append(encabezados)
        for col in range(1, len(encabezados) + 1):
            celda = ws.cell(row=1, column=col)
            celda.fill = encabezado_fill
            celda.font = encabezado_font
            celda.alignment = centrado

        for d in datos:
            ws.append([d["fecha"], d["Agropecuaria"], d["Agronegocios"], d["Agroindustrial"]])

        for col in range(1, len(encabezados) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        hojas_creadas += 1

    if hojas_creadas == 0:
        messagebox.showwarning("Exportar Excel", "No hay datos disponibles para exportar.")
        return

    wb.save(archivo)
    messagebox.showinfo("Excel", "Archivo Excel generado correctamente.")


# ──────────────────────────────────────────────────────────────────────────
# GUARDAR GRÁFICA COMO IMAGEN
# ──────────────────────────────────────────────────────────────────────────
def exportar_grafica(fig):
    """Guarda la figura de matplotlib actual (gráfica de evolución) como PNG."""
    if fig is None:
        messagebox.showwarning("Guardar gráfica", "No hay ninguna gráfica para guardar.")
        return

    archivo = filedialog.asksaveasfilename(
        defaultextension=".png",
        filetypes=[("Imagen PNG", "*.png")],
        initialfile=f"Grafica_{datetime.now().strftime('%Y%m%d_%H%M')}.png",
    )
    if not archivo:
        return

    fig.savefig(archivo, dpi=150, bbox_inches="tight", facecolor=fig.get_facecolor())
    messagebox.showinfo("Guardar gráfica", "Gráfica guardada correctamente.")