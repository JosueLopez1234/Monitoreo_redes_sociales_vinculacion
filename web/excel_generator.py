from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment


class ExcelGenerator:

    def generar(

        self,

        registros,

        total_registros,

        total_seguidores,

        total_agropecuaria,

        total_agronegocios,

        total_agroindustrial,

    ):

        wb = Workbook()

        ws = wb.active

        ws.title = "Reporte"

        verde = PatternFill(

            start_color="198754",

            end_color="198754",

            fill_type="solid",

        )

        blanco = Font(

            color="FFFFFF",

            bold=True,

        )

        negrita = Font(

            bold=True,

        )

        centro = Alignment(

            horizontal="center",

            vertical="center",

        )

        ws["A1"] = "UNIVERSIDAD LAICA ELOY ALFARO DE MANABÍ"

        ws["A2"] = "FACULTAD DE CIENCIAS AGROPECUARIAS"

        ws["A3"] = "AgroSocial Analytics"

        ws["A5"] = "Resumen General"

        for celda in [

            "A1",

            "A2",

            "A3",

            "A5",

        ]:

            ws[celda].font = negrita

        ws["A7"] = "Total Registros"

        ws["B7"] = total_registros

        ws["A8"] = "Total Seguidores"

        ws["B8"] = total_seguidores

        ws["A9"] = "Agropecuaria"

        ws["B9"] = total_agropecuaria

        ws["A10"] = "Agronegocios"

        ws["B10"] = total_agronegocios

        ws["A11"] = "Agroindustrial"

        ws["B11"] = total_agroindustrial

        fila = 14

        encabezados = [

            "Fecha",

            "Red Social",

            "Agropecuaria",

            "Agronegocios",

            "Agroindustrial",

            "Total",

        ]

        for columna, encabezado in enumerate(

            encabezados,

            start=1,

        ):

            celda = ws.cell(

                row=fila,

                column=columna,

            )

            celda.value = encabezado

            celda.fill = verde

            celda.font = blanco

            celda.alignment = centro

        fila += 1
                # ==========================================
        # AGREGAR LOS REGISTROS
        # ==========================================

        for registro in registros:

            total = (

                registro["Agropecuaria"]

                + registro["Agronegocios"]

                + registro["Agroindustrial"]

            )

            ws.cell(row=fila, column=1).value = registro["fecha"]
            ws.cell(row=fila, column=2).value = registro["red_social"]
            ws.cell(row=fila, column=3).value = registro["Agropecuaria"]
            ws.cell(row=fila, column=4).value = registro["Agronegocios"]
            ws.cell(row=fila, column=5).value = registro["Agroindustrial"]
            ws.cell(row=fila, column=6).value = total

            fila += 1

        # ==========================================
        # AJUSTAR ANCHO DE COLUMNAS
        # ==========================================

        columnas = {

            "A": 22,

            "B": 18,

            "C": 18,

            "D": 18,

            "E": 18,

            "F": 18,

        }

        for columna, ancho in columnas.items():

            ws.column_dimensions[columna].width = ancho

        # ==========================================
        # GUARDAR EN MEMORIA
        # ==========================================

        archivo = BytesIO()

        wb.save(archivo)

        archivo.seek(0)

        return archivo